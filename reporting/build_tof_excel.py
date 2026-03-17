#!/usr/bin/env python3
"""
Top of Funnel Dashboard - Excel Builder
Reads fresh HubSpot API deal data and populates the ToF ROI Analysis Excel template.
"""
import json
import copy
from openpyxl import load_workbook

# ============================================================
# CONFIG
# ============================================================

TEMPLATE_PATH = '/Users/mgottron/Downloads/Top of Funnel Dashboard.xlsx'
OUTPUT_PATH = '/Users/mgottron/Downloads/Top of Funnel Dashboard - Populated.xlsx'
SHEET_NAME = '2025 ToF ROI Analysis'

# HubSpot API result files (fresh pull Feb 17 2026)
# These 3 files contain 200 + 200 + 44 = 444 deals total
BATCH_FILES = [
    '/Users/mgottron/.claude/projects/-Users-mgottron-Claude-Code/527ec8aa-2cac-47b3-8bc8-1eefdc3a1687/tool-results/mcp-4015ff8e-8996-4ec0-b434-76a999e3f7b7-search_crm_objects-1771360380075.txt',
    '/Users/mgottron/.claude/projects/-Users-mgottron-Claude-Code/527ec8aa-2cac-47b3-8bc8-1eefdc3a1687/tool-results/mcp-4015ff8e-8996-4ec0-b434-76a999e3f7b7-search_crm_objects-1771360388902.txt',
    '/Users/mgottron/Claude Code/batch3_inline.json',
]

# Columns in the Excel (1-indexed):
# A=Team, B=ToF Activity, C=HS Subcategory
# D=SQO($), E=Closed Won($), F=CVR%, G=Still Active($), H=Pacing($), I=Pacing CVR%
# J-L=Spend cols (leave empty), M-Q=ROI cols (leave empty)
COL_SQO = 4        # D
COL_WON = 5        # E
COL_CVR = 6        # F
COL_ACTIVE = 7     # G
COL_PACING = 8     # H
COL_PACING_CVR = 9 # I

# Exclusion list: sub-categories that are NOT top-of-funnel
EXCLUDED_SUBS = {
    'marketing-additional-store', 'partnership-additional-store',
    'sales-aditionalstore', 'Customer-Reactivation', 'Customer - Upsell',
}
EXCLUDED_DEALTYPES = {'existingbusiness'}

# Won stages
WON_STAGES = {'closedwon'}
# Closed (won or lost)
CLOSED_STAGES = {'closedwon', 'closedlost'}
# Active = not closed
# "Still Active" = SQO deals that are neither won nor lost


# ============================================================
# MAPPING: dealtype → team, sub_category → Excel row
# ============================================================

# Row assignments (1-indexed, matching the Excel template)
# Marketing team rows (dealtype = PLS)
PLS_SUB_TO_ROW = {
    'Marketing-DTConnect-Matchmaking': 6,
    'Marketing-LMV-Matchmaking': 7,
    'Marketing-EXN-Matchmaking': 8,
    'Marketing-GROW-Matchmaking': 9,
    'Marketing-Events': 13,
    'Marketing-Paid-Campaign': 15,
    'PLS-Other': 18,          # → Mass Emailing (Senders)
    'Marketing-Outbound': 19, # → Mass Emailing (Contractor)
    'Marketing-Free-trial': 21,
}

# Partnership team rows (dealtype = Partnership)
def partnership_row(sub):
    if not sub:
        return None
    sub_lower = sub.lower()
    if 'event' in sub_lower:
        return 12  # Partnerships Events
    if sub == 'Partnerships-Referral':
        return 24
    if sub == 'Partnership-Free-trial':
        return 24  # Grouped with Partnerships-Referral
    return None

# Sales team rows (dealtype = Outbound Outreach)
OUTBOUND_SUB_TO_ROW = {
    'Outbound - Cold Calling': 27,
    'Sales-HappyStack': 28,
    'Outbound - Senders': 27,  # → BDR Internal (keep under Sales since dealtype=Outbound Outreach)
    'Sales - AE Generated': 29,
}


def map_deal_to_row(dealtype, sub_category):
    """Map a deal to its Excel row based on dealtype (team) and sub_category (channel)."""
    dt = (dealtype or '').strip()
    sub = (sub_category or '').strip()

    # Exclude non-ToF
    if sub.lower() in {s.lower() for s in EXCLUDED_SUBS}:
        return None
    if dt.lower() in {s.lower() for s in EXCLUDED_DEALTYPES}:
        return None

    if dt == 'PLS':
        return PLS_SUB_TO_ROW.get(sub)
    elif dt == 'Partnership':
        return partnership_row(sub)
    elif dt == 'Outbound Outreach':
        return OUTBOUND_SUB_TO_ROW.get(sub, 27)  # Default to BDR Internal
    else:
        return None


# Rollup definitions: parent_row → list of child rows to sum
ROLLUPS = {
    5:  [6, 7, 8, 9],           # Affiliate Partners
    11: [12, 13],                # Events Consolidated
    14: [15],                    # Paid Advertising
    17: [18, 19],                # Email Marketing
    20: [21],                    # Inbound
    4:  [6, 7, 8, 9, 13, 15, 18, 19, 21],  # Marketing Total (excludes P-Events row 12)
    23: [24],                    # Partnerships Total (Events roll up under Events, not here)
    26: [27, 28],                # BDRs
    25: [27, 28, 29],            # Sales Total
}


# ============================================================
# DATA LOADING
# ============================================================

def load_deals_from_file(filepath):
    """Load deals from a HubSpot API result file (MCP wrapper format)."""
    with open(filepath, 'r') as f:
        raw = json.load(f)
    inner = json.loads(raw[0]['text'])
    return inner.get('results', [])


def load_deals_inline(deals_json):
    """Load deals from inline JSON results."""
    return deals_json


def parse_amount(val):
    if val is None or val == '' or val == 'null':
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


# ============================================================
# MAIN
# ============================================================

def main():
    # Load all deals from batch files
    all_raw_deals = []
    for fp in BATCH_FILES:
        batch = load_deals_from_file(fp)
        print(f"Loaded {len(batch)} deals from {fp.split('/')[-1]}")
        all_raw_deals.extend(batch)

    # Batch 3 was returned inline (44 deals) - load from the last API call
    # These are the remaining deals (offset 400+)
    # We already have them in the third API response above, let's load that too
    batch3_file = '/Users/mgottron/.claude/projects/-Users-mgottron-Claude-Code/527ec8aa-2cac-47b3-8bc8-1eefdc3a1687/tool-results/mcp-4015ff8e-8996-4ec0-b434-76a999e3f7b7-search_crm_objects-1771360388902.txt'
    # Actually batch 3 was returned directly in the conversation (44 deals at offset 400)
    # But those deals might also be in batch file 2 if it contains offset 200-399
    # Let me check: batch 1 = offset 0-199 (200 deals), batch 2 = offset 200-399 (200 deals)
    # The 44 deals at offset 400 were returned inline in the API response
    # Let me just use the inline data from the most recent pull

    # Deduplicate by deal ID
    deals_by_id = {}
    for raw in all_raw_deals:
        deal_id = str(raw.get('id', ''))
        if deal_id:
            deals_by_id[deal_id] = raw

    print(f"\nTotal unique deals from batch files: {len(deals_by_id)}")

    print(f"\nTotal unique deals from ALL result files: {len(deals_by_id)}")

    # Filter: createdate in 2025, has SQO date
    filtered_deals = {}
    for deal_id, raw in deals_by_id.items():
        props = raw.get('properties', {})
        createdate = props.get('createdate', '')
        sqo_date = props.get('hs_v2_date_entered_presentationscheduled', '')
        if createdate and createdate.startswith('2025') and sqo_date:
            filtered_deals[deal_id] = raw

    print(f"Deals created in 2025 with SQO date: {len(filtered_deals)}")

    # Process deals into row buckets
    # For each row, track: sqo_total, won_total, active_total
    row_data = {}  # row_number → {sqo: float, won: float, active: float}
    unmapped = []

    for deal_id, raw in filtered_deals.items():
        props = raw.get('properties', {})
        dealtype = props.get('dealtype', '')
        sub = props.get('deal_type___sub_category', '')
        amount = parse_amount(props.get('amount'))
        stage = props.get('dealstage', '')
        dealname = props.get('dealname', '')

        row = map_deal_to_row(dealtype, sub)
        if row is None:
            unmapped.append({
                'id': deal_id,
                'name': dealname,
                'dealtype': dealtype,
                'sub': sub,
                'amount': amount,
                'stage': stage,
            })
            continue

        if row not in row_data:
            row_data[row] = {'sqo': 0.0, 'won': 0.0, 'active': 0.0}

        # Every deal in our set reached SQO
        row_data[row]['sqo'] += amount

        if stage in WON_STAGES:
            row_data[row]['won'] += amount
        elif stage not in CLOSED_STAGES:
            # Still active (not won, not lost)
            row_data[row]['active'] += amount
        # If closedlost: counted in SQO but not in won or active

    # Print unmapped deals
    if unmapped:
        print(f"\n--- Unmapped deals ({len(unmapped)}) ---")
        for d in unmapped:
            print(f"  {d['name']}: dealtype={d['dealtype']}, sub={d['sub']}, "
                  f"amount=${d['amount']:.2f}, stage={d['stage']}")

    # Print row-level data
    ROW_NAMES = {
        6: 'DTConnect', 7: 'Landmark', 8: 'EXN', 9: 'GROW',
        12: 'Partnerships Events', 13: 'Marketing Events',
        15: 'LinkedIn Paid Ads',
        18: 'Mass Emailing (Senders)', 19: 'Mass Emailing (Contractor)',
        21: 'Inbound Leads',
        24: 'Partnerships-Referral',
        27: 'Outbound-BDRs (Internal)', 28: 'Outbound-BDRs (External)',
        29: 'AE-Generated',
    }

    print(f"\n--- Row-level data ---")
    for row in sorted(row_data.keys()):
        d = row_data[row]
        name = ROW_NAMES.get(row, f'Row {row}')
        print(f"  Row {row:2d} ({name:30s}): SQO=${d['sqo']:>10,.2f}  "
              f"Won=${d['won']:>10,.2f}  Active=${d['active']:>10,.2f}")

    # Compute rollups
    for parent_row, child_rows in ROLLUPS.items():
        parent = {'sqo': 0.0, 'won': 0.0, 'active': 0.0}
        for cr in child_rows:
            if cr in row_data:
                parent['sqo'] += row_data[cr]['sqo']
                parent['won'] += row_data[cr]['won']
                parent['active'] += row_data[cr]['active']
        row_data[parent_row] = parent

    ROLLUP_NAMES = {
        4: 'MARKETING TOTAL', 5: 'Affiliate Partners', 11: 'Events Consolidated',
        14: 'Paid Advertising', 17: 'Email Marketing', 20: 'Inbound',
        23: 'PARTNERSHIPS TOTAL', 25: 'SALES TOTAL', 26: 'BDRs',
    }

    print(f"\n--- Rollup totals ---")
    for row in sorted(ROLLUP_NAMES.keys()):
        if row in row_data:
            d = row_data[row]
            name = ROLLUP_NAMES[row]
            print(f"  Row {row:2d} ({name:25s}): SQO=${d['sqo']:>10,.2f}  "
                  f"Won=${d['won']:>10,.2f}  Active=${d['active']:>10,.2f}")

    # Team breakdown for verification
    mkt_won = row_data.get(4, {}).get('won', 0)
    prt_won = row_data.get(23, {}).get('won', 0)
    sales_won = row_data.get(25, {}).get('won', 0)
    print(f"\n--- Team Won totals ---")
    print(f"  Marketing:    ${mkt_won:>10,.2f}")
    print(f"  Partnerships: ${prt_won:>10,.2f}")
    print(f"  Sales:        ${sales_won:>10,.2f}")
    print(f"  GRAND TOTAL:  ${mkt_won + prt_won + sales_won:>10,.2f}")

    # ============================================================
    # WRITE TO EXCEL
    # ============================================================
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[SHEET_NAME]

    # Number formats
    DOLLAR_FMT = '"$"#,##0'
    PCT_FMT = '0.0%'

    # Leaf rows: write raw data values + cell formulas for CVR/Pacing
    LEAF_ROWS = [6, 7, 8, 9, 12, 13, 15, 18, 19, 21, 24, 27, 28, 29]

    for row_num in LEAF_ROWS:
        d = row_data.get(row_num)
        if d is None:
            # No data for this row — leave blank
            continue

        sqo = d['sqo']
        won = d['won']
        active = d['active']
        r = row_num

        # D: SQO ($) — raw value
        cell = ws.cell(row=r, column=COL_SQO, value=round(sqo, 2))
        cell.number_format = DOLLAR_FMT

        # E: Won ($) — raw value
        cell = ws.cell(row=r, column=COL_WON, value=round(won, 2))
        cell.number_format = DOLLAR_FMT

        # F: CVR% — formula =E/D (as decimal, formatted as %)
        cell = ws.cell(row=r, column=COL_CVR)
        cell.value = f'=IF(D{r}=0,"",E{r}/D{r})'
        cell.number_format = PCT_FMT

        # G: Still Active ($) — raw value
        cell = ws.cell(row=r, column=COL_ACTIVE, value=round(active, 2))
        cell.number_format = DOLLAR_FMT

        # H: Pacing ($) — formula =E + G * (E/D)
        cell = ws.cell(row=r, column=COL_PACING)
        cell.value = f'=IF(D{r}=0,"",E{r}+(G{r}*(E{r}/D{r})))'
        cell.number_format = DOLLAR_FMT

        # I: Pacing CVR% — formula =H/D (as decimal, formatted as %)
        cell = ws.cell(row=r, column=COL_PACING_CVR)
        cell.value = f'=IF(D{r}=0,"",H{r}/D{r})'
        cell.number_format = PCT_FMT

    # Rollup rows: write SUM/formula references
    # Template already has formulas for rows 4 and 5 — fix them to use proper CVR
    # Row 24 template formulas are broken (reference row 25=Sales) — overwrite above as leaf

    # Rollup definitions: parent_row → (child row refs for SUM)
    ROLLUP_FORMULAS = {
        # Row 4 and 5 already have SUM formulas in the template for D,E,G,H
        # But their CVR% uses AVERAGE which is wrong — fix those
        11: ('D12+D13', 'E12+E13', 'G12+G13', 'H12+H13'),
        14: ('D15', 'E15', 'G15', 'H15'),
        17: ('D18+D19', 'E18+E19', 'G18+G19', 'H18+H19'),
        20: ('D21', 'E21', 'G21', 'H21'),
        23: ('D24', 'E24', 'G24', 'H24'),
        26: ('D27+D28', 'E27+E28', 'G27+G28', 'H27+H28'),
        25: ('D27+D28+D29', 'E27+E28+E29', 'G27+G28+G29', 'H27+H28+H29'),
    }

    for r, (d_ref, e_ref, g_ref, h_ref) in ROLLUP_FORMULAS.items():
        # D: SQO
        cell = ws.cell(row=r, column=COL_SQO)
        cell.value = f'={d_ref}'
        cell.number_format = DOLLAR_FMT

        # E: Won
        cell = ws.cell(row=r, column=COL_WON)
        cell.value = f'={e_ref}'
        cell.number_format = DOLLAR_FMT

        # F: CVR% = Won / SQO (dollar-weighted, not average)
        cell = ws.cell(row=r, column=COL_CVR)
        cell.value = f'=IF(D{r}=0,"",E{r}/D{r})'
        cell.number_format = PCT_FMT

        # G: Active
        cell = ws.cell(row=r, column=COL_ACTIVE)
        cell.value = f'={g_ref}'
        cell.number_format = DOLLAR_FMT

        # H: Pacing
        cell = ws.cell(row=r, column=COL_PACING)
        cell.value = f'={h_ref}'
        cell.number_format = DOLLAR_FMT

        # I: Pacing CVR%
        cell = ws.cell(row=r, column=COL_PACING_CVR)
        cell.value = f'=IF(D{r}=0,"",H{r}/D{r})'
        cell.number_format = PCT_FMT

    # Fix rows 4 and 5 — template has formulas but CVR% uses AVERAGE (wrong)
    # Overwrite only the CVR% and Pacing CVR% columns; leave D,E,G,H formulas intact
    for r in [4, 5]:
        # Fix F: CVR% — dollar-weighted
        cell = ws.cell(row=r, column=COL_CVR)
        cell.value = f'=IF(D{r}=0,"",E{r}/D{r})'
        cell.number_format = PCT_FMT

        # Fix I: Pacing CVR%
        cell = ws.cell(row=r, column=COL_PACING_CVR)
        cell.value = f'=IF(D{r}=0,"",H{r}/D{r})'
        cell.number_format = PCT_FMT

        # Standardize number formats on the existing formula cells
        ws.cell(row=r, column=COL_SQO).number_format = DOLLAR_FMT
        ws.cell(row=r, column=COL_WON).number_format = DOLLAR_FMT
        ws.cell(row=r, column=COL_ACTIVE).number_format = DOLLAR_FMT
        ws.cell(row=r, column=COL_PACING).number_format = DOLLAR_FMT

    wb.save(OUTPUT_PATH)
    print(f"\nExcel saved to: {OUTPUT_PATH}")


if __name__ == '__main__':
    main()
